import time
from openpyxl import Workbook
import re
from selenium import webdriver
import time
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.chrome.service import Service as ChromeService
from selenium.webdriver.common.by import By


def run():
    book = Workbook()
    sheet = book.active
    error_sheet = book.create_sheet("errors")


    sheet.cell(row=1, column=1).value = "Nom et prenom"
    sheet.cell(row=1, column=2).value = "Date de naissance"
    sheet.cell(row=1, column=3).value = "Taille (en cm)"
    sheet.cell(row=1, column=4).value = "Poid (en kg)"
    sheet.cell(row=1, column=5).value = "Poste"
    sheet.cell(row=1, column=6).value = "Club actuel"

    error_sheet.cell(row=1, column=1).value = "Page URL"
    error_sheet.cell(row=1, column=2).value = "Error"

    i = 2
    j = 2

    service = ChromeService(executable_path=ChromeDriverManager().install())

    for link in links:
        print(link)
        
        options = webdriver.ChromeOptions() 
        options.add_argument('--headless')
        errors = []
        
        with webdriver.Chrome(service=service, options=options) as driver:  
            driver.get(link)

            try:
                name = driver.find_element(By.XPATH,'//h1[@id="page-title"]')
                sheet.cell(row=i, column=1).value = name.text
            except Exception as e:
                sheet.cell(row=i, column=1).value = ""

                error_sheet.cell(row=j, column=1).value = link
                error_sheet.cell(row=j, column=2).value = e.__str__()
                j += 1
            try:
                # With regex
                birthdate = driver.find_element(By.XPATH,'//span[contains(., "Date")]/following-sibling::span')
                sheet.cell(row=i, column=2).value = birthdate.text
            except Exception as e:
                sheet.cell(row=i, column=2).value = ""

                error_sheet.cell(row=j, column=1).value = link
                error_sheet.cell(row=j, column=2).value = e.__str__()
                j += 1
            try:
                # With regex
                height = driver.find_element(By.XPATH,'//span[contains(., "Taille")]/following-sibling::span')
                sheet.cell(row=i, column=3).value = int(height.text.split(" ")[0])
            except Exception as e:
                sheet.cell(row=i, column=3).value = ""

                error_sheet.cell(row=j, column=1).value = link
                error_sheet.cell(row=j, column=2).value = e.__str__()
                j += 1
            try:
                # With regex
                weight = driver.find_element(By.XPATH,'//span[contains(., "Poids")]/following-sibling::span')
                sheet.cell(row=i, column=4).value = int(weight.text.split(" ")[0])
            except Exception as e:
                sheet.cell(row=i, column=4).value = ""

                error_sheet.cell(row=j, column=1).value = link
                error_sheet.cell(row=j, column=2).value = e.__str__()
                j += 1
            try:
                position = driver.find_element(By.XPATH,'//div[@class="visu-infos"]/h2')
                sheet.cell(row=i, column=5).value = position.text
            except Exception as e:
                sheet.cell(row=i, column=5).value = ""

                error_sheet.cell(row=j, column=1).value = link
                error_sheet.cell(row=j, column=2).value = e.__str__()
                j += 1
            try:
                team = driver.find_element(By.XPATH,'//ul[@class="infos-list"]/li/h3')
                sheet.cell(row=i, column=6).value = team.text
            except Exception as e:
                sheet.cell(row=i, column=6).value = ""

            i += 1
            time.sleep(2)

    book.save("liste_joueurs_top_14.xlsx")